import os
import time

from tkinter import *
import tkinter.ttk as ttk
from tkinter.filedialog import askdirectory, asksaveasfilename

import handling

import openpyxl


ABC = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z',
       'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL', 'AM', 'AN', 'AO', 'AP', 'AQ', 'AR', 'AS', 'AT', 'AU', 'AV', 'AW', 'AX', 'AY', 'AZ',
       'BA', 'BB', 'BC', 'BD', 'BE', 'BF', 'BG', 'BH', 'BI', 'BJ', 'BK', 'BL', 'BM', 'BN', 'BO', 'BP', 'BQ', 'BR', 'BS', 'BT', 'BU', 'BV', 'BW', 'BX', 'BY', 'BZ',
       'CA', 'CB', 'CC', 'CD', 'CE', 'CF', 'CG', 'CH', 'CI', 'CJ', 'CK', 'CL', 'CM', 'CN', 'CO', 'CP', 'CQ', 'CR', 'CS', 'CT', 'CU', 'CV', 'CW', 'CX', 'CY', 'CZ',
       'DA', 'DB', 'DC', 'DD', 'DE', 'DF', 'DG', 'DH', 'DI', 'DJ', 'DK', 'DL', 'DM', 'DN', 'DO', 'DP', 'DQ', 'DR', 'DS', 'DT', 'DU', 'DV', 'DW', 'DX', 'DY', 'DZ',
       'EA', 'EB', 'EC', 'ED', 'EE', 'EF', 'EG', 'EH', 'EI', 'EJ', 'EK', 'EL', 'EM', 'EN', 'EO', 'EP', 'EQ', 'ER', 'ES', 'ET', 'EU', 'EV', 'EW', 'EX', 'EY', 'EZ',
       'FA', 'FB', 'FC', 'FD', 'FE', 'FF', 'FG', 'FH', 'FI', 'FJ', 'FK', 'FL', 'FM', 'FN', 'FO', 'FP', 'FQ', 'FR', 'FS', 'FT', 'FU', 'FV', 'FW', 'FX', 'FY', 'FZ',
       'GA', 'GB', 'GC', 'GD', 'GE', 'GF', 'GG', 'GH', 'GI', 'GJ', 'GK', 'GL', 'GM', 'GN', 'GO', 'GP', 'GQ', 'GR', 'GS', 'GT', 'GU', 'GV', 'GW', 'GX', 'GY', 'GZ',
       'HA', 'HB', 'HC', 'HD', 'HE', 'HF', 'HG', 'HH', 'HI', 'HJ', 'HK', 'HL', 'HM', 'HN', 'HO', 'HP', 'HQ', 'HR', 'HS', 'HT', 'HU', 'HV', 'HW', 'HX', 'HY', 'HZ',
       'IA', 'IB', 'IC', 'ID', 'IE', 'IF', 'IG', 'IH', 'II', 'IJ', 'IK', 'IL', 'IM', 'IN', 'IO', 'IP', 'IQ', 'IR', 'IS', 'IT', 'IU', 'IV', 'IW', 'IX', 'IY', 'IZ',
       'JA', 'JB', 'JC', 'JD', 'JE', 'JF', 'JG', 'JH', 'JI', 'JJ', 'JK', 'JL', 'JM', 'JN', 'JO', 'JP', 'JQ', 'JR', 'JS', 'JT', 'JU', 'JV', 'JW', 'JX', 'JY', 'JZ',
       'KA', 'KB', 'KC', 'KD', 'KE', 'KF', 'KG', 'KH', 'KI', 'KJ', 'KK', 'KL', 'KM', 'KN', 'KO', 'KP', 'KQ', 'KR', 'KS', 'KT', 'KU', 'KV', 'KW', 'KX', 'KY', 'KZ',
       'LA', 'LB', 'LC', 'LD', 'LE', 'LF', 'LG', 'LH', 'LI', 'LJ', 'LK', 'LL', 'LM', 'LN', 'LO', 'LP', 'LQ', 'LR', 'LS', 'LT', 'LU', 'LV', 'LW', 'LX', 'LY', 'LZ',
       'MA', 'MB', 'MC', 'MD', 'ME', 'MF', 'MG', 'MH', 'MI', 'MJ', 'MK', 'ML', 'MM', 'MN', 'MO', 'MP', 'MQ', 'MR', 'MS', 'MT', 'MU', 'MV', 'MW', 'MX', 'MY', 'MZ',
       'NA', 'NB', 'NC', 'ND', 'NE', 'NF', 'NG', 'NH', 'NI', 'NJ', 'NK', 'NL', 'NM', 'NN', 'NO', 'NP', 'NQ', 'NR', 'NS', 'NT', 'NU', 'NV', 'NW', 'NX', 'NY', 'NZ',
       'OA', 'OB', 'OC', 'OD', 'OE', 'OF', 'OG', 'OH', 'OI', 'OJ', 'OK', 'OL', 'OM', 'ON', 'OO', 'OP', 'OQ', 'OR', 'OS', 'OT', 'OU', 'OV', 'OW', 'OX', 'OY', 'OZ',
       'PA', 'PB', 'PC', 'PD', 'PE', 'PF', 'PG', 'PH', 'PI', 'PJ', 'PK', 'PL', 'PM', 'PN', 'PO', 'PP', 'PQ', 'PR', 'PS', 'PT', 'PU', 'PV', 'PW', 'PX', 'PY', 'PZ',
       'QA', 'QB', 'QC', 'QD', 'QE', 'QF', 'QG', 'QH', 'QI', 'QJ', 'QK', 'QL', 'QM', 'QN', 'QO', 'QP', 'QQ', 'QR', 'QS', 'QT', 'QU', 'QV', 'QW', 'QX', 'QY', 'QZ',
       'RA', 'RB', 'RC', 'RD', 'RE', 'RF', 'RG', 'RH', 'RI', 'RJ', 'RK', 'RL', 'RM', 'RN', 'RO', 'RP', 'RQ', 'RR', 'RS', 'RT', 'RU', 'RV', 'RW', 'RX', 'RY', 'RZ',
       'SA', 'SB', 'SC', 'SD', 'SE', 'SF', 'SG', 'SH', 'SI', 'SJ', 'SK', 'SL', 'SM', 'SN', 'SO', 'SP', 'SQ', 'SR', 'SS', 'ST', 'SU', 'SV', 'SW', 'SX', 'SY', 'SZ',
       'TA', 'TB', 'TC', 'TD', 'TE', 'TF', 'TG', 'TH', 'TI', 'TJ', 'TK', 'TL', 'TM', 'TN', 'TO', 'TP', 'TQ', 'TR', 'TS', 'TT', 'TU', 'TV', 'TW', 'TX', 'TY', 'TZ',
       'UA', 'UB', 'UC', 'UD', 'UE', 'UF', 'UG', 'UH', 'UI', 'UJ', 'UK', 'UL', 'UM', 'UN', 'UO', 'UP', 'UQ', 'UR', 'US', 'UT', 'UU', 'UV', 'UW', 'UX', 'UY', 'UZ',
       'VA', 'VB', 'VC', 'VD', 'VE', 'VF', 'VG', 'VH', 'VI', 'VJ', 'VK', 'VL', 'VM', 'VN', 'VO', 'VP', 'VQ', 'VR', 'VS', 'VT', 'VU', 'VV', 'VW', 'VX', 'VY', 'VZ',
       'WA', 'WB', 'WC', 'WD', 'WE', 'WF', 'WG', 'WH', 'WI', 'WJ', 'WK', 'WL', 'WM', 'WN', 'WO', 'WP', 'WQ', 'WR', 'WS', 'WT', 'WU', 'WV', 'WW', 'WX', 'WY', 'WZ',
       'XA', 'XB', 'XC', 'XD', 'XE', 'XF', 'XG', 'XH', 'XI', 'XJ', 'XK', 'XL', 'XM', 'XN', 'XO', 'XP', 'XQ', 'XR', 'XS', 'XT', 'XU', 'XV', 'XW', 'XX', 'XY', 'XZ',
       'YA', 'YB', 'YC', 'YD', 'YE', 'YF', 'YG', 'YH', 'YI', 'YJ', 'YK', 'YL', 'YM', 'YN', 'YO', 'YP', 'YQ', 'YR', 'YS', 'YT', 'YU', 'YV', 'YW', 'YX', 'YY', 'YZ',
       'ZA', 'ZB', 'ZC', 'ZD', 'ZE', 'ZF', 'ZG', 'ZH', 'ZI', 'ZJ', 'ZK', 'ZL', 'ZM', 'ZN', 'ZO', 'ZP', 'ZQ', 'ZR', 'ZS', 'ZT', 'ZU', 'ZV', 'ZW', 'ZX', 'ZY', 'ZZ']
path = ''
balls = {}
right_answers = set()
stop = False

root = Tk()
main_menu = Menu(root)
root.config(menu=main_menu)

def open_folder(*args):
    global stop
    global balls
    save_right_answers()
    path = askdirectory()
    if not path: return
    start_time = time.time()
    all_files = len(list(os.scandir(path)))
    dir_data = os.scandir(path)
    pb['length'] = 200
    n = 0
    for item in dir_data:
        if stop:
            stop = False
            message_string['text'] = 'Операция прервана'
            message_string['bg'] = 'yellow'
            return
        if item.is_file():
            filepath = item.path.replace('/', os.path.sep)
            try:
                result = handling.handle(filepath)
            except Exception:
                message_string['text'] = f'Ошибка в файле {filepath}'
                message_string['bg'] = 'red'
                return
            balls[result[0]] = (len(right_answers.intersection(result[2])), result[1], sorted([i[0]+' '+i[1] for i in result[2]]))
        else:
            message_string['text'] = f'Внимание, в папке не только изображения'
            message_string['bg'] = 'orange'
        pb.step(100/all_files)
        pb.update()
        message_string['text'] = f'Обработка... {round(n*100/all_files)}%'
        n += 1
    message_string['text'] = f'Обработано! {round(time.time()-start_time, 2)} c'

def save_file(*args):
    if not balls:
        message_string['text'] = 'Бланки ещё не обработаны!'
        message_string['bg'] = 'red'
        return
    message_string['bg'] = 'green'
    message_string['text'] = 'Сохранение...'
    
    results_file = openpyxl.Workbook()
    ws = results_file.active
    ws.title = "Результаты"
    ws['A1'].value = "Имя"
    ws['B1'].value = "Фамилия"
    ws['C1'].value = "Баллы"
    ws['D1'].value = "Класс"
    for i in range(len(balls[list(balls)[0]][2])):
        ws[ABC[i+4]+'1'].value = i+1
    
    for n, i in enumerate(balls):
        ws['A'+str(n+2)].value = i.split(' ')[0]
        ws['B'+str(n+2)].value = i.split(' ')[1]
        ws['C'+str(n+2)].value = balls[i][0]
        ws['D'+str(n+2)].value = balls[i][1]
        for j in range(len(balls[i][2])):
            ws[ABC[j+4]+str(n+2)].value = balls[i][2][j].split(' ')[1]
    path = asksaveasfilename(filetypes=[("Excel files", ".xlsx")])
    if not path: return
    if not path.endswith('.xlsx'): path += '.xlsx'
    if os.path.exists(path):
        try:
            os.remove(path)
        except PermissionError:
            message_string['text'] = 'Закройте файл в другой программе'
            message_string['bg'] = 'red'
            
    results_file.save(path)
    message_string['bg'] = 'green'
    message_string['text'] = 'Файл успешно сохранён'

def save_right_answers(*args):
    global right_answers
    right_answers = set()
    try:
        for i in right_answers_input.get(0.0, END).split('\n'):
            if i.strip():
                if len(i.strip()) < 2: raise ValueError('invalid syntax of answers!') #проверка на правильность синтаксиса
                right_answers.add((i.strip()[0], i.strip()[1:]))
        message_string['bg'] = 'green'
        message_string['text'] = 'All ok'
    except Exception:
        message_string['text'] = 'Неправильный синтаксис ответов!'
        message_string['bg'] = 'red'

def cancel_operation(*args):
    global stop
    stop = True

message_string = Label(root, text='')
message_string.pack(side=TOP)

pb = ttk.Progressbar(root, length=200)
pb.pack(side=TOP, pady=5)

cancel_button = Button(text='Cancel', bg='orange', command=cancel_operation)
cancel_button.pack(side=TOP)
root.bind_all('<Control-c>', cancel_operation)

file_menu = Menu(main_menu, tearoff=0)
file_menu.add_command(label="Open (ctrl-o)", command=open_folder)
root.bind_all('<Control-o>', open_folder)
file_menu.add_command(label="Save as (ctrl-s)", command=save_file)
root.bind_all('<Control-s>', save_file)
main_menu.add_cascade(label='File', menu=file_menu)

right_answers_input_frame = Frame(root, bg='gray')
right_answers_input_frame.pack(side=TOP, pady=50)


right_answers_input_label = Label(right_answers_input_frame, text='Вводите правильные ответы сюда:', bg='gray')
right_answers_input_label.pack(side=TOP)

right_answers_input = Text(right_answers_input_frame, width=20, height=15)
right_answers_input.pack(side=TOP, padx=60)
right_answers_input.bind('<Return>', save_right_answers)

root.mainloop()
